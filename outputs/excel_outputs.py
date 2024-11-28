import openpyxl
from openpyxl.utils import get_column_letter

def save_to_excel(data, filename="output.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    # إضافة العناوين في الصف الأول
    headers = ["ID", "Question", "Answer", "Model Used", "Timestamp"]
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f"{col_letter}1"] = header

    # إضافة البيانات
    for row_num, record in enumerate(data, 2):
        for col_num, value in enumerate(record, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}{row_num}"] = value

    # حفظ الملف
    wb.save(filename)
    return filename
